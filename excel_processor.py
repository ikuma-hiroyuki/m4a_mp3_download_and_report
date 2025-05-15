from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
import utils


class ExcelProcessor:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.sheets = []

    def load_excel(self):
        """Excelファイルを読み込み、シート情報を取得する"""
        try:
            # ファイルが開かれているかチェック
            if utils.is_excel_file_open(self.excel_file_path):
                return False, "Excelファイルが開かれています。閉じてから処理を実行してください。"

            # Excelファイルを読み込む
            self.workbook = load_workbook(self.excel_file_path)
            self.sheets = self.workbook.sheetnames
            return True, None
        except Exception as e:
            return False, f"Excelファイル読み込みエラー: {str(e)}"

    def get_sheets(self):
        """利用可能なシート名のリストを返す"""
        return self.sheets

    def get_links_from_sheets(self, target_sheets, check_l_column=True):
        """指定されたシートからリンクを収集する

        Args:
            target_sheets: 処理対象のシート名リスト
            check_l_column: Trueの場合、L列に値がある行をスキップする
        """
        links_data = []

        for sheet_name in target_sheets:
            if sheet_name in self.sheets:
                sheet = self.workbook[sheet_name]
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):  # ヘッダー行をスキップ
                    if len(row) >= 10:  # J列（10列目）が存在する場合
                        cell_j = row[9]  # J列（インデックスは0から始まるので9）
                        cell_value = cell_j.value

                        # L列の値をチェック
                        cell_l = row[11]
                        cell_l_value = cell_l.value
                        if check_l_column and cell_l_value:
                            continue  # L列に値がある場合はスキップ

                        # セルの値がURLかどうかチェック
                        if cell_value and isinstance(cell_value, str) and "drive.google.com/file" in cell_value:
                            links_data.append({
                                'sheet_name': sheet_name,
                                'row_num': row_idx,
                                'url': cell_value
                            })
                        # ハイパーリンクがある場合
                        elif cell_j.hyperlink and cell_j.hyperlink.target and "drive.google.com/file" in cell_j.hyperlink.target:
                            links_data.append({
                                'sheet_name': sheet_name,
                                'row_num': row_idx,
                                'url': cell_j.hyperlink.target
                            })

        return links_data

    def update_cell_with_hyperlink(self, sheet_name, row_num, file_name, url, duration=None):
        """セルをハイパーリンク付きテキストに更新する"""
        sheet = self.workbook[sheet_name]
        cell = sheet.cell(row=row_num, column=10)  # J列

        # ハイパーリンクを設定
        cell.value = file_name
        # Hyperlinkオブジェクトを直接作成
        cell.hyperlink = Hyperlink(target=url, ref=cell.coordinate, tooltip=f"Open {file_name}")

        # 再生時間が提供されている場合はL列に出力
        if duration:
            duration_cell = sheet.cell(row=row_num, column=12)  # L列
            duration_cell.value = duration

    def save_results_to_sheet(self, results_df):
        """結果を新しいシートに保存する"""
        if not results_df.empty:
            # 既存のresultsシートを削除（存在する場合）
            if 'results' in self.workbook.sheetnames:
                del self.workbook['results']

            # 新しいresultsシートを作成
            results_sheet = self.workbook.create_sheet('results')

            # ヘッダーを設定
            headers = results_df.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                results_sheet.cell(row=1, column=col_idx).value = header

            # データを書き込み
            for row_idx, row in enumerate(results_df.itertuples(), 2):
                for col_idx, value in enumerate(row[1:], 1):
                    results_sheet.cell(row=row_idx, column=col_idx).value = value

            # 列幅を自動調整（簡易的な実装）
            for col_idx in range(1, len(headers) + 1):
                results_sheet.column_dimensions[chr(64 + col_idx)].width = 20

    def save_excel(self):
        """Excelファイルを保存する"""
        try:
            self.workbook.save(self.excel_file_path)
            return True, None
        except Exception as e:
            return False, f"Excelファイル保存エラー: {str(e)}"

    def save_results(self, pl_df):
        """Polarsを使用して結果をresultsシートに保存する"""
        try:
            # resultsシートの取得または作成
            if 'results' in self.workbook.sheetnames:
                # 既存のシートを取得
                results_sheet = self.workbook['results']
                print("既存のresultsシートを上書きします")
            else:
                # 新しいシートを作成
                results_sheet = self.workbook.create_sheet('results')
                print("新しいresultsシートを作成しました")

            # ヘッダーを設定
            headers = pl_df.columns
            for col_idx, header in enumerate(headers, 1):
                results_sheet.cell(row=1, column=col_idx).value = header

            # Polarsデータフレームをリストに変換して一気に書き込み
            data_rows = pl_df.rows()
            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    results_sheet.cell(row=row_idx, column=col_idx).value = value

            # 列幅を自動調整
            for col_idx in range(1, len(headers) + 1):
                results_sheet.column_dimensions[chr(64 + col_idx)].width = 20

            print(f"{len(data_rows)}行のデータを書き込みました")
            return True, None
        except Exception as e:
            print(f"結果シート作成エラー: {str(e)}")
            return False, f"結果保存エラー: {str(e)}"

    def open_excel(self):
        """処理後にExcelファイルを開く"""
        utils.open_excel_file(self.excel_file_path)
